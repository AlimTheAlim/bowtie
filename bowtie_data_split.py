"""
bowtie_data_split.py

Reads labeling from Excel (.xlsx) files and copies images into:
OUT/{train,val,test}/{accept,reject}/

Fixes included:
- Skips junk Excel files starting with '.' or '~$'
- Verifies real .xlsx by ZIP signature (PK)
- Wraps load_workbook with try/except (won't crash)
- Ignores hidden image files starting with '.'
- Maps image ID -> zero-padded filename like 003.jpg
- Stratified split 70/15/15 by accept/reject

EDIT the CONFIG section then run:
    python bowtie_data_split.py
"""

import os
import re
import shutil
import random
from pathlib import Path
from zipfile import BadZipFile

import openpyxl

# ===================== CONFIG (EDIT THESE) =====================
ROOT = Path(r"D:\CSCI-484\bowtie\LA-UR-25-28525-bowtie\round-2")   # contains .xlsx + image folders
OUT  = Path(r"D:\CSCI-484\bowtie\LA-UR-25-28525-bowtie\dataset")   # output dataset folder

# Column letters in your sheet:
COL_IMG_ID = "A"     # image id column (e.g., 3, 14, 116 OR "003")
COL_AR     = "C"     # column that contains A/R (accept/reject) for that image
HEADER_ROW = 2       # row index where data starts (1 if no header)

# Which Excel sheet(s) to read:
SHEETS = None        # None = all sheets, OR ["Sheet1"] etc.

# Folder mapping:
# Option 1 (default): folder name == Excel filename stem (e.g., 116.xlsx -> folder 116_zoomed_in ???)
# We'll try these in order:
FOLDER_CANDIDATES = [
    "{sheet}",        # folder == sheet name
    "{xlsx_stem}",    # folder == excel filename without extension
]

# If your real folders are like "116_zoomed_in" and excel is "116.xlsx",
# you can add patterns here:
FOLDER_GLOB_PATTERNS = [
    "{xlsx_stem}*",    # e.g., 116* -> 116_zoomed_in, 116_zoomed_out
]

# Splits
TRAIN_PCT, VAL_PCT, TEST_PCT = 0.70, 0.15, 0.15
SEED = 42

# Image extensions to match
IMG_EXTS = [".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp"]
# ===============================================================


def is_real_xlsx(path: Path) -> bool:
    """Real .xlsx files are zip containers (start with 'PK')."""
    try:
        with open(path, "rb") as f:
            sig = f.read(2)
        return sig == b"PK"
    except OSError:
        return False


def normalize_id(v) -> str | None:
    """
    Convert cell value to filename base like '003' or '116'.
    - numeric: int -> zero-pad to 3 digits if < 1000
    - string: extract first digit run -> same rules
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        n = int(v)
        return f"{n:03d}" if n < 1000 else str(n)

    s = str(v).strip()
    m = re.search(r"\d+", s)
    if not m:
        return None
    n = int(m.group(0))
    return f"{n:03d}" if n < 1000 else str(n)


def label_from_ar(v) -> str | None:
    """Map A/R to accept/reject."""
    if v is None:
        return None
    s = str(v).strip().upper()
    if s == "A":
        return "accept"
    if s == "R":
        return "reject"
    return None


def resolve_image_folder(root: Path, xlsx: Path, sheet_name: str) -> Path | None:
    """Find best matching folder for a given sheet/xlsx using candidates + glob patterns."""
    # direct candidates
    for templ in FOLDER_CANDIDATES:
        folder_name = templ.format(sheet=sheet_name, xlsx_stem=xlsx.stem)
        p = root / folder_name
        if p.exists() and p.is_dir():
            return p

    # glob patterns
    for g in FOLDER_GLOB_PATTERNS:
        pat = g.format(sheet=sheet_name, xlsx_stem=xlsx.stem)
        hits = sorted([p for p in root.glob(pat) if p.is_dir() and not p.name.startswith(".")])
        if len(hits) == 1:
            return hits[0]
        # if multiple, prefer the one that contains at least one image
        for h in hits:
            for ext in IMG_EXTS:
                if any(h.glob(f"*{ext}")):
                    return h
    return None


def find_image(folder: Path, base: str) -> Path | None:
    """Locate an image file in folder for a given id base (e.g., '003'). Skips hidden dotfiles."""
    # exact matches first
    for ext in IMG_EXTS:
        p = folder / f"{base}{ext}"
        if p.exists() and p.is_file() and not p.name.startswith("."):
            return p

    # fallback: startswith base
    for p in folder.iterdir():
        if p.name.startswith("."):
            continue
        if p.is_file() and p.suffix.lower() in IMG_EXTS and p.stem.startswith(base):
            return p

    return None


def stratified_split(items, train_pct, val_pct, seed=42):
    """
    items: list of (path, label)
    returns train, val, test with stratification by label
    """
    random.seed(seed)
    by_label = {}
    for x in items:
        by_label.setdefault(x[1], []).append(x)

    train, val, test = [], [], []
    for lbl, lst in by_label.items():
        random.shuffle(lst)
        n = len(lst)

        n_train = int(round(train_pct * n))
        n_val = int(round(val_pct * n))

        # guard against rounding overflow
        if n_train + n_val > n:
            n_val = max(0, n - n_train)

        train += lst[:n_train]
        val += lst[n_train:n_train + n_val]
        test += lst[n_train + n_val:]

    random.shuffle(train)
    random.shuffle(val)
    random.shuffle(test)
    return train, val, test


def safe_copy(src: Path, dst_dir: Path):
    """Copy src into dst_dir avoiding collisions."""
    dst_dir.mkdir(parents=True, exist_ok=True)
    dst = dst_dir / src.name
    if dst.exists():
        dst = dst_dir / f"{src.stem}__{src.parent.name}{src.suffix}"
    shutil.copy2(src, dst)


def main():
    if abs((TRAIN_PCT + VAL_PCT + TEST_PCT) - 1.0) > 1e-9:
        raise ValueError("TRAIN/VAL/TEST must sum to 1.0")

    OUT.mkdir(parents=True, exist_ok=True)

    # Discover valid .xlsx files
    xlsx_files = []
    for f in ROOT.glob("*.xlsx"):
        name = f.name
        if name.startswith(".") or name.startswith("~$"):
            continue
        if not is_real_xlsx(f):
            print(f"[SKIP] not a real xlsx (zip signature missing): {name}")
            continue
        xlsx_files.append(f)
    xlsx_files = sorted(xlsx_files)

    if not xlsx_files:
        raise FileNotFoundError(f"No valid .xlsx files found in {ROOT}")

    samples = []
    missing_images = []
    skipped_workbooks = 0
    total_rows_seen = 0

    for xlsx in xlsx_files:
        try:
            wb = openpyxl.load_workbook(xlsx, data_only=True)
        except BadZipFile as e:
            print(f"[SKIP] bad xlsx zip: {xlsx.name} ({e})")
            skipped_workbooks += 1
            continue
        except Exception as e:
            print(f"[SKIP] cannot open workbook: {xlsx.name} ({e})")
            skipped_workbooks += 1
            continue

        sheet_names = wb.sheetnames if SHEETS is None else SHEETS

        for sh_name in sheet_names:
            if sh_name not in wb.sheetnames:
                continue
            sh = wb[sh_name]

            img_folder = resolve_image_folder(ROOT, xlsx, sh_name)
            if img_folder is None:
                print(f"[WARN] No image folder match for {xlsx.name} / sheet '{sh_name}'. Skipping sheet.")
                continue

            row = HEADER_ROW
            empty_run = 0

            while True:
                img_id = sh[f"{COL_IMG_ID}{row}"].value
                ar_val = sh[f"{COL_AR}{row}"].value

                if img_id is None and ar_val is None:
                    empty_run += 1
                    if empty_run >= 5:
                        break
                    row += 1
                    continue
                empty_run = 0

                base = normalize_id(img_id)
                lbl = label_from_ar(ar_val)

                total_rows_seen += 1

                if base and lbl:
                    p = find_image(img_folder, base)
                    if p is None:
                        missing_images.append((xlsx.name, sh_name, row, img_folder.name, base))
                    else:
                        samples.append((p, lbl))

                row += 1

    if not samples:
        raise RuntimeError(
            "No samples collected.\n"
            "Check COL_IMG_ID/COL_AR/HEADER_ROW and folder mapping rules."
        )

    # Split
    train, val, test = stratified_split(samples, TRAIN_PCT, VAL_PCT, seed=SEED)

    # Create + copy
    for split_name, split_items in [("train", train), ("val", val), ("test", test)]:
        for (src, lbl) in split_items:
            safe_copy(src, OUT / split_name / lbl)

    # Reports
    (OUT / "reports").mkdir(parents=True, exist_ok=True)

    # missing images report
    if missing_images:
        rep = OUT / "reports" / "missing_images.tsv"
        with rep.open("w", encoding="utf-8") as f:
            f.write("xlsx\tsheet\trow\tfolder\timg_id\n")
            for x in missing_images:
                f.write("\t".join(map(str, x)) + "\n")

    # summary report
    def count_label(lst, label):
        return sum(1 for _, l in lst if l == label)

    summary = OUT / "reports" / "summary.txt"
    with summary.open("w", encoding="utf-8") as f:
        f.write(f"ROOT: {ROOT}\n")
        f.write(f"OUT:  {OUT}\n\n")
        f.write(f"Valid .xlsx files: {len(xlsx_files)}\n")
        f.write(f"Skipped workbooks: {skipped_workbooks}\n")
        f.write(f"Total rows scanned (approx): {total_rows_seen}\n")
        f.write(f"Samples collected: {len(samples)}\n")
        f.write(f"Missing images: {len(missing_images)}\n\n")

        f.write("Split counts:\n")
        f.write(f"  Train: {len(train)} (accept={count_label(train,'accept')}, reject={count_label(train,'reject')})\n")
        f.write(f"  Val:   {len(val)} (accept={count_label(val,'accept')}, reject={count_label(val,'reject')})\n")
        f.write(f"  Test:  {len(test)} (accept={count_label(test,'accept')}, reject={count_label(test,'reject')})\n")

    print("========================================")
    print(f"Collected samples: {len(samples)}")
    print(f"Missing images:    {len(missing_images)}")
    print(f"Train/Val/Test:    {len(train)}/{len(val)}/{len(test)}")
    print(f"Reports in:        {OUT / 'reports'}")
    print("Done.")


if __name__ == "__main__":
    main()